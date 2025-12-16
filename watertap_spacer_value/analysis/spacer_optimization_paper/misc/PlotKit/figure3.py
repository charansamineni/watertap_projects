import numpy as np
from numpy.core.defchararray import capitalize
from psPlotKit.data_plotter.fig_generator import figureGenerator
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from psPlotKit.data_plotter.ps_box_plotter import boxPlotter
from watertap_spacer_value.plotting_tools.paul_tol_color_maps import gen_paultol_colormap



# 1. Use lowercase Tol name
colormap_name = "tol_bright"

# 2. Generate color array
my_colors = gen_paultol_colormap(colormap_name, num_samples=5, return_map=False)


swro_file = "../../spacer_scale_benchmarks/literature_correlations/SWRO_results.xlsx"
wb = load_workbook(swro_file, read_only=True)
sheet_names = wb.sheetnames
print(sheet_names)
for sheet_name in sheet_names:
    if 'macro' in sheet_name:
        swro_macros = pd.read_excel(swro_file, sheet_name=sheet_name)



        # 3. Assign to figureGenerator the expected interface
        fig = figureGenerator(
            colormap=colormap_name,  # pass string key, not Colormap object
        )

        # 4. Make sure the colors are stored in fig.colorMaps
        fig.colorMaps[colormap_name] = my_colors
        fig.colormaps = colormap_name
        fig.init_figure(
            fig_size=(4, 4),
            fig_dpi=300,
            nrows=1,
            ncols=1,
            sharex=False,
            sharey=False,
            tight_layout=True,
            save_location="",
            save_folder=None,
            save_name="Figure_3_SWRO_LCOW_comparison",
            show_figs=True,
        )
        fig.set_default_figure_settings(
            font_size=6, label_size= 8
        )
        corr_list = swro_macros['correlation'].tolist()
        corr_label = [c.capitalize() for c in corr_list]
        lcow_data = swro_macros["LCOW"].tolist()
        for i, corr in enumerate(corr_list):
            fig.plot_bar(
            i, lcow_data[i],
            xerr = None,
            yerr = None,
            bottom = None,
            width = 0.4,
            edgecolor = "black",
            color = None,
            align = "center",
            ax_idx = 0,
            label = None,
            vertical = True,
            linewidth = 1,
            save_label = None,
            zorder = 4,
            ecolor = "black",
            capsize = 4,
            )
        fig.set_axis_ticklabels(xticklabels=corr_label,
                                ylabel="LCOW (USD/m3)",
                                ylims=(0, 12),
                                rotate=True,
                                angle=0,
                                fontsize=8,
                                ha="center",
                                va="top")
        fig.show()







